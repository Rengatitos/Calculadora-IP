import tkinter as tk
from tkinter import ttk, messagebox
import ipaddress
from conversiones import ipv4_to_ipv6, ipv6_to_ipv4, calcular_subredes
from validaciones import validate_ip
import os

subredes = []

def launch_gui():
    tipo_subneteo_combo = None  

    def crear_contenido_pestana(pestana):
        
       if pestana == tab_vlsm:
                input_frame = ttk.Frame(pestana)
                input_frame.pack(fill=tk.X, pady=5)

                tk.Label(
                    input_frame,
                    text="Ingresa una dirección IP (IPv4 o IPv6):",
                    font=("Arial", 12, "bold"),
                    bg="#FF5733"
                ).grid(row=0, column=0, padx=10, pady=5, sticky="w")
                ip_entry = tk.Entry(
                    input_frame,
                    width=35,
                    bg="#900C3F",      
                    fg="white",        
                    insertbackground="white",
                    relief="solid",
                    bd=1,
                    font=("Arial", 12, "bold") 
                )
                ip_entry.grid(row=0, column=1, padx=10, pady=5,ipady=5)
                
                tk.Label(
                    input_frame,
                    text="Tipo de Subneteo:",
                    font=("Arial", 12, "bold"),
                    bg="#FF5733",
                ).grid(row=1, column=0, padx=10, pady=5, sticky="w")
                tipo_subneteo_combo = tk.StringVar()
                
                def option_changed(*args):
                    if (tipo_subneteo_combo.get()=="Proporcional (fijo)"):
                        #mostrar_hosts_entry.config(state='disabled')
                        hosts_entry.config(state='disabled')
                        cantidad_entry.config(state='normal')
                        
                    else:
                        #mostrar_hosts_entry.config(state='normal')
                        hosts_entry.config(state='normal')
                        cantidad_entry.config(state='disabled')
                        
                    print(tipo_subneteo_combo.get())

                combo = tk.OptionMenu(input_frame, tipo_subneteo_combo, " ","Desproporcional (VLSM)", "Proporcional (fijo)", command=option_changed)
                combo.config(
                    width=31,
                    bg="#900C3F",
                    fg="white",
                    activebackground="#A93256",
                    activeforeground="white",
                    highlightthickness=1,
                    relief="solid",
                    font=("Arial", 12, "bold"),
                )
                combo.grid(row=1, column=1, padx=10, pady=5)
                tipo_subneteo_combo.set(" ")
                
                tk.Label(
                    input_frame,
                    text="Número de subred",
                    font=("Arial", 12, "bold"),
                    bg="#FF5733"
                ).grid(row=1, column=2, columnspan=2, padx=10, pady=5, sticky="w")
                cantidad_entry = tk.Entry(
                    input_frame,
                    width=35,
                    bg="#900C3F",         # Fondo personalizado
                    fg="white",           # Texto blanco
                    insertbackground="white",  # Cursor blanco
                    relief="solid",
                    bd=1,
                    font=("Arial", 12, "bold") 

                )
                cantidad_entry.grid(row=1, column=5, columnspan=2, padx=10, pady=5, ipady=10, sticky="w")
                
                
                tk.Label(
                    input_frame,
                    text="Máscara",
                    font=("Arial", 12, "bold"),
                    bg="#FF5733"
                ).grid(row=2, column=2, columnspan=2, padx=10, pady=5, sticky="w")
                mascara_entry = tk.Entry(
                    input_frame,
                    width=35,
                    bg="#900C3F",         # Fondo personalizado
                    fg="white",           # Texto blanco
                    insertbackground="white",  # Cursor blanco
                    relief="solid",
                    bd=1,
                    font=("Arial", 12, "bold") 

                )
                mascara_entry.grid(row=2, column=5, columnspan=2, padx=10, pady=5, ipady=10, sticky="w")
                

                tk.Label(
                    input_frame,
                    text="Hosts por subred (separados por coma):",
                    font=("Arial", 12, "bold"),
                    bg="#FF5733"
                ).grid(row=2, column=0, padx=10, pady=5, sticky="w")
                hosts_entry = tk.Entry(
                    input_frame,
                    width=35,
                    bg="#900C3F",         # Fondo personalizado
                    fg="white",           # Texto blanco
                    insertbackground="white",  # Cursor blanco
                    relief="solid",
                    bd=1,
                    font=("Arial", 12, "bold") 

                )
                hosts_entry.grid(row=2, column=1, padx=10, pady=5, ipady=10)
                
                


                tk.Label(
                    input_frame,
                    text="Cantidad de hosts a mostrar:",
                    font=("Arial", 12, "bold"),
                    bg="#FF5733"
                ).grid(row=3, column=0, padx=10, pady=5, sticky="w")
                mostrar_hosts_entry = tk.Entry(
                    input_frame,
                    width=35,
                    bg="#900C3F",           # Fondo vino
                    fg="white",             # Texto blanco
                    insertbackground="white",  # Cursor blanco
                    relief="solid",
                    bd=1,
                    font=("Arial", 12, "bold") 
                )
                mostrar_hosts_entry.grid(row=3, column=1, padx=10, pady=5, ipady=5)
                mostrar_hosts_entry.insert(0, " ")

                

                ip_type_label = tk.Label(
                    pestana,
                    text="Tipo de IP: ",
                    font=("Arial", 12,"bold"),
                    bg="#FF5733",
                    fg="black"
                )
                ip_type_label.pack(pady=5)


                # Marco con borde amarillo
                result_frame = tk.Frame(
                    pestana,
                    width=1400,
                    height=60,
                    bg="#900C3F",        # Color del borde
                    highlightthickness=0,
                    bd=2,               # Grosor del borde
                    relief="solid",
                )
                result_frame.pack_propagate(False)
                result_frame.pack(pady=10)

                result_label = ttk.Label(
                    result_frame,
                    text="",
                    justify=tk.LEFT,
                    wraplength=400,
                    font=("Arial", 12,"bold"),
                    background="#FF5733"
                )
                result_label.pack(padx=10, pady=5)



                copy_button = ttk.Button(pestana, text="Copiar Resultado", state=tk.DISABLED)
                copy_button.pack(pady=10)

                
                # Bloque fijo con fondo naranja (#FF5733)
                white_block = tk.Frame(pestana, width=1400, height=200, relief="solid", bg="#FF5733", bd=1)
                white_block.pack_propagate(False)
                white_block.pack(pady=10)

                # Crear el Canvas primero
                tabla_canvas = tk.Canvas(white_block, width=1400, height=200, bg="#900C3F", highlightthickness=0)

                # Scroll horizontal
                tabla_scroll_x = tk.Scrollbar(white_block, orient=tk.HORIZONTAL, command=tabla_canvas.xview)
                tabla_canvas.configure(xscrollcommand=tabla_scroll_x.set)

                # Scroll vertical
                tabla_scroll_y = tk.Scrollbar(white_block, orient=tk.VERTICAL, command=tabla_canvas.yview)
                tabla_canvas.configure(yscrollcommand=tabla_scroll_y.set)

                # Empacar los scrollbars
                tabla_scroll_x.pack(side=tk.BOTTOM, fill=tk.X)
                tabla_scroll_y.pack(side=tk.RIGHT, fill=tk.Y)

                # Empacar el canvas
                tabla_canvas.pack(side=tk.TOP, fill=tk.BOTH, expand=True, padx=10, pady=10)

                # Crear el frame dentro del Canvas para agregar contenido
                tabla_frame = tk.Frame(tabla_canvas, bg="#900C3F")
                tabla_canvas.create_window((0, 0), window=tabla_frame, anchor="nw")

                # Ajustar scroll dinámico al tamaño de la tabla
                def ajustar_scroll(event):
                    tabla_canvas.configure(scrollregion=tabla_canvas.bbox("all"))

                tabla_frame.bind("<Configure>", ajustar_scroll)



                download_subnet_button = ttk.Button(
                    input_frame,
                    text="Descargar Informe Sub redes",
                    state=tk.DISABLED,
                )
                download_subnet_button.grid(row=0, column=5, padx=10, pady=5)

                                
                from PIL import Image, ImageTk  # Asegúrate de tener Pillow instalado (pip install Pillow)

                def ver_topologia_logica():
                    global subredes

                    if not subredes:
                        messagebox.showwarning("Advertencia", "Primero genera las subredes.")
                        return

                    ventana_topologia = tk.Toplevel()
                    ventana_topologia.title("Topología lógica de subredes")
                    ventana_topologia.geometry("1200x600")

                    canvas = tk.Canvas(ventana_topologia, bg="#655EAA", scrollregion=(0, 0, 2500, 1000))
                    canvas.pack(fill=tk.BOTH, expand=True)

                    scrollbar_y = tk.Scrollbar(ventana_topologia, orient=tk.VERTICAL, command=canvas.yview)
                    scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)
                    canvas.config(yscrollcommand=scrollbar_y.set)

                    # Coordenadas
                    origen_x = 600
                    origen_y = 100
                    espacio_horizontal = 300
                    base_y = origen_y + 150
                    num_subredes = len(subredes)

                    # Cargar y mostrar imagen del router principal
                    try:
                        router_img = Image.open("router.jpg")
                        router_img = router_img.resize((60, 60), resample=Image.LANCZOS)
                        router_icon = ImageTk.PhotoImage(router_img)

                        canvas.create_image(origen_x, origen_y, image=router_icon)
                        canvas.router_icon = router_icon  # mantener referencia
                        canvas.create_text(origen_x, origen_y - 40, text="Router Principal", font=("Arial", 10, "bold"))
                    except Exception as e:
                        print("⚠️ Error cargando imagen:", e)
                        messagebox.showerror("Error", f"No se pudo cargar la imagen del router:\n{e}")
                        canvas.create_oval(origen_x - 30, origen_y - 30, origen_x + 30, origen_y + 30, fill="#DDAEA3")
                        canvas.create_text(origen_x, origen_y, text="Router\nPrincipal", font=("Arial", 10, "bold"))
                    # Cargar imagen del ordenador una sola vez
                    try:
                        ordenador_img = Image.open("ordenador.png")
                        ordenador_img = ordenador_img.resize((40, 40), resample=Image.LANCZOS)
                        ordenador_icon = ImageTk.PhotoImage(ordenador_img)
                    except Exception as e:
                        print("⚠️ Error cargando ordenador.png:", e)
                        ordenador_icon = None

                    # Para evitar que se borren las imágenes múltiples
                    canvas.ordenador_icons = []
                    # Dibujar cada subred conectada
                    for i, sub in enumerate(subredes):
                        offset = (i - (num_subredes - 1) / 2) * espacio_horizontal
                        x = origen_x + offset
                        y = base_y
                        canvas.create_line(origen_x, origen_y + 60, x, y - 30, arrow=tk.LAST, fill="#555", width=2)
                        if ordenador_icon:
                            canvas.create_image(x, y - 80, image=ordenador_icon)
                            canvas.ordenador_icons.append(ordenador_icon)  # mantener referencia
                        # subredes
                        canvas.create_rectangle(x - 130, y - 40, x + 130, y + 90, fill="#7878C7", outline="#333", width=2)
                        canvas.create_text(x, y - 15, text=f"Subred {i+1}", font=("Arial", 10, "bold"), fill="white")
                        canvas.create_text(x, y + 5, text=f"{sub['subred']} ({sub['mascara']})", font=("Arial", 9), fill="white")
                        canvas.create_text(x, y + 25, text=f"Switch: {sub['router']}", font=("Arial", 9), fill="white")
                        canvas.create_text(x, y + 45, text=f"Broadcast: {sub['broadcast']}", font=("Arial", 9), fill="white")

                def convert_ip():
                    ip_input = ip_entry.get().strip()
                    if not ip_input:
                        messagebox.showerror("Error", "Por favor ingresa una dirección IP")
                        return
                    ip_type = validate_ip(ip_input)
                    if not ip_type:
                        messagebox.showerror("Error", "La dirección IP no es válida.")
                        return
                    ip_type_label.config(text=f"Tipo de IP: {ip_type}")
                    if ip_type == "IPv4":
                        ipv6_result = ipv4_to_ipv6(ip_input)
                        if ipv6_result:
                            result_label.config(text=f"IPv4 original: {ip_input}\nConvertido a IPv6: {ipv6_result}")
                            copy_button.config(state=tk.NORMAL)
                    elif ip_type == "IPv6":
                        ipv4_result = ipv6_to_ipv4(ip_input)
                        if ipv4_result:
                            result_label.config(text=f"IPv6 original: {ip_input}\nConvertido a IPv4: {ipv4_result}")
                            copy_button.config(state=tk.NORMAL)
               
                def copy_to_clipboard():
                    root.clipboard_clear()
                    root.clipboard_append(result_label.cget("text"))
                    messagebox.showinfo("Copiado", "El resultado ha sido copiado al portapapeles")
                
                def deshabilitar_campos():
                    tipo_subneteo=tipo_subneteo = tipo_subneteo_combo.get()
                    print(tipo_subneteo)

                def ver_subredes():
                    global subredes
                    ip_base = ip_entry.get().strip()
                    lista_hosts = hosts_entry.get().strip()
                    tipo_subneteo = tipo_subneteo_combo.get()
                

                    # Validación de IP
                    try:
                        red_obj = ipaddress.IPv4Network(f"{ip_base}/24", strict=False)
                    except ValueError:
                        messagebox.showerror("Error", "La IP base no es válida. Ejemplo: 192.168.1.0")
                        return

                    total_ips_disponibles = red_obj.num_addresses

                    # Validación de lista de hosts
                    try:
                        hosts_por_subred = [int(x.strip()) for x in lista_hosts.split(",") if x.strip()]
                    except ValueError:
                        messagebox.showerror("Error", "Por favor ingresa una lista válida de cantidades de hosts (ej: 50,10,5).")
                        return

                    # Validación de cantidad de hosts a mostrar
                    try:
                        cantidad_a_mostrar = int(mostrar_hosts_entry.get().strip())
                    except ValueError:
                        messagebox.showerror("Error", "Ingresa una cantidad válida de hosts a mostrar.")
                        return

                    subredes = []
                    
                    #-------------------------------------------------------------
                    try:
                        entrada_mascara = mascara_entry.get().strip()

                        if "/" in entrada_mascara:
                            entrada_mascara = entrada_mascara.replace("/", "")

                        if "." in entrada_mascara:
                            # Máscara decimal como 255.255.0.0
                            red_obj = ipaddress.IPv4Network(f"{ip_base}/{entrada_mascara}", strict=False)
                            prefijo_usuario = red_obj.prefixlen
                        else:
                            # Prefijo CIDR como 16
                            prefijo_usuario = int(entrada_mascara)
                            if not (0 <= prefijo_usuario <= 32):
                                raise ValueError
                            red_obj = ipaddress.IPv4Network(f"{ip_base}/{prefijo_usuario}", strict=False)

                        # Ajustamos IP base real con la red calculada
                        ip_base = str(red_obj.network_address)

                        # Usamos IP base con prefijo (lo que espera calcular_subredes)
                        ip_cidr_base = f"{ip_base}/{prefijo_usuario}"

                    except Exception:
                        messagebox.showerror("Error", "Máscara inválida. Usa 255.255.0.0 o 24.")
                        return
          
                    #------------------------------------------------


                    # Subneteo proporcional (VLSM)
                    # Subneteo desproporcional (VLSM)
                    if tipo_subneteo == "Desproporcional (VLSM)":
                        total_requerido = sum([(h + 2) for h in hosts_por_subred])
                        if total_requerido > total_ips_disponibles:
                            messagebox.showerror(
                                "Error",
                                f"No es posible realizar el subneteo VLSM con los valores dados.\n"
                                f"Se requieren al menos {total_requerido} direcciones IP, pero la red base solo tiene {total_ips_disponibles}."
                            )
                            return

                        #
                        subredes = calcular_subredes(ip_cidr_base, hosts_por_subred, cantidad_a_mostrar)

                    
                    # Subneteo desproporcional (fijo)
                    # Subneteo proporcional (fijo)
                    elif tipo_subneteo == "Proporcional (fijo)":
                        try:
                            cantidad_subredes = int(cantidad_entry.get().strip())
                            
                            if cantidad_subredes <= 0:
                                raise ValueError
                        except ValueError:
                            messagebox.showerror("Error", "Cantidad de subredes debe ser un número entero positivo.")
                            return
                        if cantidad_subredes & (cantidad_subredes - 1) != 0:
                            messagebox.showerror("Error", "La cantidad de subredes debe ser una potencia de 2 (1, 2, 4, 8, 16, etc.).")
                            return

    


                        bits_necesarios = (cantidad_subredes - 1).bit_length()
                        prefijo = red_obj.prefixlen + bits_necesarios
                       #------------------------------------ 
                        if red_obj.prefixlen + bits_necesarios > 32:
                            messagebox.showerror("Error", "No hay suficientes bits para dividir la red en esa cantidad de subredes.")
                            return
                        #------------------------------------ 
                        try:
                            subredes_disponibles = list(red_obj.subnets(new_prefix=prefijo))
                        except ValueError:
                            messagebox.showerror("Error", "No se pueden generar las subredes con la cantidad especificada.")
                            return

                        if len(subredes_disponibles) < cantidad_subredes:
                            messagebox.showerror("Error", f"No hay suficientes subredes disponibles. Solo se pueden crear {len(subredes_disponibles)}.")
                            return

                        subredes = []
                        for nueva_red in subredes_disponibles[:cantidad_subredes]:
                            hosts = list(nueva_red.hosts())
                            router = hosts[0] if hosts else None
                            hosts_mostrados = hosts[1:1 + cantidad_a_mostrar] if len(hosts) > 1 else []
                            ultimo = hosts[-1] if hosts else None

                            datos = {
                                "subred": str(nueva_red),
                                "mascara": str(nueva_red.netmask),
                                "prefijo": nueva_red.prefixlen,
                                "cantidad_hosts": nueva_red.num_addresses - 2,
                                "inicio": str(nueva_red.network_address),
                                "fin": str(nueva_red.broadcast_address),
                                "router": str(router) if router else None,
                                "broadcast": str(nueva_red.broadcast_address),
                                "ultimo_host": str(ultimo) if ultimo else None
                            }

                            for i, h in enumerate(hosts_mostrados):
                                datos[f"host_{i + 1}"] = str(h)

                            subredes.append(datos)


                    # Mostrar resultados en la tabla
                    for widget in tabla_frame.winfo_children():
                        widget.destroy()

                    headers = [
                        "SubRed", "Máscara de red", "Tamaño", "Cantidad de Hosts disponibles", "Dirección Inicial",
                        "Dirección Final", "Dirección Router", "Dirección Broadcast"
                    ] + [f"Añadir a host {i+1}" for i in range(cantidad_a_mostrar)] + ["Última dirección host"]

                    for col, header in enumerate(headers):
                        tk.Label(tabla_frame, text=header, font=("Arial", 10, "bold"), bg="white").grid(row=0, column=col, padx=5, pady=2)

                    for i, sub in enumerate(subredes, start=1):
                        values = [
                            i,
                            str(sub["mascara"]),
                            f"/{sub['prefijo']}",
                            sub["cantidad_hosts"],
                            str(sub["inicio"]),
                            str(sub["fin"]),
                            str(sub["router"]),
                            str(sub["broadcast"])
                        ]

                        for j in range(1, cantidad_a_mostrar + 1):
                            values.append(str(sub.get(f"host_{j}", "-")))

                        values.append(str(sub["ultimo_host"]))

                        for col, val in enumerate(values):
                            tk.Label(tabla_frame, text=val, font=("Arial", 10), bg="white").grid(row=i, column=col, padx=5, pady=2)

                    download_subnet_button.config(state=tk.NORMAL)


                def descargar_subredes():
                    import pandas as pd
                    import os
                    from openpyxl import load_workbook, Workbook
                    from openpyxl.styles import Font, Alignment, PatternFill
                    from openpyxl.utils.dataframe import dataframe_to_rows
                    from openpyxl.utils import get_column_letter
                    from datetime import datetime
                    from tkinter import filedialog

                    # Verificamos que haya datos
                    children = tabla_frame.winfo_children()
                    if not children:
                        messagebox.showwarning("Advertencia", "No hay datos para exportar.")
                        return

                    # Determinar número de columnas
                    columnas = 0
                    for widget in children:
                        info = widget.grid_info()
                        columnas = max(columnas, info["column"] + 1)

                    # Leer el contenido de la tabla en orden por fila y columna
                    filas_ordenadas = sorted(children, key=lambda w: (w.grid_info()["row"], w.grid_info()["column"]))
                    contenido = []
                    fila_actual = -1
                    fila = []
                    for widget in filas_ordenadas:
                        text = widget.cget("text")
                        row = widget.grid_info()["row"]
                        if row != fila_actual:
                            if fila:
                                contenido.append(fila)
                            fila = []
                            fila_actual = row
                        fila.append(text)
                    if fila:
                        contenido.append(fila)

                    # Validamos contenido
                    if len(contenido) < 2:
                        messagebox.showwarning("Advertencia", "No hay suficientes datos para exportar.")
                        return

                    headers = contenido[0]
                    data = contenido[1:]
                    df = pd.DataFrame(data, columns=headers)

                    # Información extra
                    ip_base_valor = ip_entry.get().strip()
                    hosts_lista_valor = hosts_entry.get().strip()
                    tipo = tipo_subneteo_combo.get()
                    hoja_nombre = tipo.split()[0] + "_" + datetime.now().strftime("%H%M%S")

                    extra_info = pd.DataFrame([
                        ["IP Base", ip_base_valor],
                        ["Hosts por subred", hosts_lista_valor],
                        ["Tipo Subneteo", tipo]
                    ], columns=["Detalle", "Valor"])

                    # Diálogo para guardar archivo
                    nombre_sugerido = "informe_subredes.xlsx"
                    ruta_archivo = filedialog.asksaveasfilename(
                        initialfile=nombre_sugerido,
                        defaultextension=".xlsx",
                        filetypes=[("Excel files", "*.xlsx")],
                        title="Guardar informe de subredes como..."
                    )

                    if not ruta_archivo:
                        return

                    try:
                        if os.path.exists(ruta_archivo):
                            wb = load_workbook(ruta_archivo)
                        else:
                            wb = Workbook()
                            del wb[wb.sheetnames[0]]

                        ws = wb.create_sheet(title=hoja_nombre)

                        # Escribir información extra
                        for r_idx, row in enumerate(dataframe_to_rows(extra_info, index=False, header=True), 1):
                            for c_idx, value in enumerate(row, 1):
                                ws.cell(row=r_idx, column=c_idx, value=value)

                        # Escribir tabla
                        tabla_start_row = len(extra_info) + 3
                        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), tabla_start_row):
                            for c_idx, value in enumerate(row, 1):
                                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                                if r_idx == tabla_start_row:
                                    cell.font = Font(bold=True, color="FFFFFF")
                                    cell.fill = PatternFill("solid", fgColor="4F81BD")
                                cell.alignment = Alignment(horizontal="center", vertical="center")

                        # Autoajuste columnas
                        for col in ws.columns:
                            max_length = 0
                            col_letter = get_column_letter(col[0].column)
                            for cell in col:
                                if cell.value:
                                    max_length = max(max_length, len(str(cell.value)))
                            ws.column_dimensions[col_letter].width = max_length + 2

                        ws.freeze_panes = f"A{tabla_start_row + 1}"
                        ws.auto_filter.ref = ws.dimensions

                        wb.save(ruta_archivo)
                        messagebox.showinfo("Éxito", f"Informe guardado correctamente en hoja '{hoja_nombre}'.")

                    except Exception as e:
                        messagebox.showerror("Error", f"No se pudo guardar el archivo:\n{e}")

                def limpiar_campos():
                    ip_entry.delete(0, tk.END)
                    hosts_entry.delete(0, tk.END)
                    mostrar_hosts_entry.delete(0, tk.END)
                    mostrar_hosts_entry.insert(0, " ")
                    result_label.config(text="")
                    cantidad_entry.delete(0,tk.END)
                    cantidad_entry.insert(0,"  ")
                    mascara_entry.delete(0,tk.END)
                    mascara_entry.insert(0," ")
                    ip_type_label.config(text="Tipo de IP: ")
                    copy_button.config(state=tk.DISABLED)
                    tipo_subneteo_combo.set(" ")
                    download_subnet_button.config(state=tk.DISABLED)
                    
                    for widget in tabla_frame.winfo_children():
                        widget.destroy()
                button_topologia = ttk.Button(
                    input_frame,
                    text="Ver topologia",
                    command=ver_topologia_logica,
                    style="Yellow.TButton"
                )
                button_topologia.grid(row=0, column=6, padx=10, pady=5)
                convertir_button = ttk.Button(
                    input_frame,
                    text="Convertir",
                    command=convert_ip,
                    style="Yellow.TButton"
                )
                convertir_button.grid(row=0, column=2, padx=10, pady=5)
                ttk.Button(input_frame, text="Ver Sub redes", command=ver_subredes).grid(row=0, column=3, padx=10, pady=5)
                # Botón Limpiar (fuera de input_frame, debajo del white_block)
                limpiar_button = ttk.Button(pestana, text="Limpiar", command=limpiar_campos)
                limpiar_button.pack(pady=(0, 10))  # espacio arriba opcional

                limpiar_button.pack(pady=(0, 10))
                limpiar_button.pack_configure(anchor="center")

                copy_button.config(command=copy_to_clipboard)
                download_subnet_button.config(command=descargar_subredes)



#######################################################
    # Inicio GUI

    root = tk.Tk()
    root.title("Conversor Avanzado de IPs")
    root.state('zoomed')
    root.configure(bg='#FF5733')

    style = ttk.Style()
    style.configure("TFrame", background="#581845")
    
    style.configure("TLabel", background="#E0BBE4", font=("Arial", 12))
    style.configure("TButton", font=("Arial", 10), padding=10, relief="flat", background="#FF5733")
    style.configure("Header.TLabel", font=("Arial", 14, "bold"), foreground="#333333")
    style.configure("Result.TFrame", relief=tk.SUNKEN, background="#D1C4E9", padding=15)
    style.configure("TEntry", font=("Arial", 12), padding=5)
    style.configure("WhiteBlock.TFrame", background="white")

    main_frame = ttk.Frame(root, padding="20")
    main_frame.pack(fill=tk.BOTH, expand=True)

    title_label = tk.Label(
    main_frame,
    text="Conversor de IPs",
    font=("Arial", 14, "bold"),
    bg="#FF5733",
    fg="black"
    )
    title_label.pack(pady=(20, 10))

    estilo = ttk.Style()


    estilo.configure("Estilo.TNotebook", background="#C70039", borderwidth=0)

    estilo.configure("Estilo.TNotebook.Tab",
        background="black",
        foreground="#FF5733",
        font=("Arial", 10, "bold"),
        padding=3
    )

    estilo.map("Estilo.TNotebook.Tab",
        background=[("selected", "black")],
        foreground=[("selected", "#FF5733")]
    )

    notebook = ttk.Notebook(main_frame, style="Estilo.TNotebook")
    notebook.pack(fill=tk.BOTH, expand=True)

    tab_vlsm = ttk.Frame(notebook)
    tab_prop = ttk.Frame(notebook)
    notebook.add(tab_vlsm, text='Convertidor')
    notebook.add(tab_prop, text='Historial')

    crear_contenido_pestana(tab_vlsm)
    crear_contenido_pestana(tab_prop)

    root.mainloop()
